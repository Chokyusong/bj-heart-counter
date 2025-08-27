# app.py
import io, re, csv, zipfile, unicodedata
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="BJ별 하트 정리 — 관리자용 & BJ용", layout="wide")
st.title("BJ별 하트 정리 자동화")
st.caption("CSV/XLSX 업로드 → 참여BJ별 개별 엑셀 파일 생성 (좌: 관리자용, 우: BJ용)  / 모든 처리 메모리 전용")

uploaded = st.file_uploader("CSV 또는 엑셀(.xlsx) 업로드", type=["csv", "xlsx"])
sheet_name = st.text_input("시트 이름 (엑셀일 때만, 비우면 첫 시트)", value="")

# -------------------- autosize helpers --------------------
def visual_len(val) -> int:
    """한글/전각/이모지를 넓게 계산하는 표시 폭 길이."""
    s = str(val) if val is not None else ""
    w = 0
    for ch in s:
        # east_asian_width: F/W/A(전각/넓은/모호)는 2칸
        if unicodedata.east_asian_width(ch) in ("F", "W", "A"):
            w += 2
        elif ord(ch) >= 0x1F300:  # 이모지 대략 영역
            w += 2
        else:
            w += 1
    return w

def autosize_columns(wb, min_w=12, max_w=80, pad=2):
    for ws in wb.worksheets:
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_width = 0
            for cell in col:
                if cell.value is not None:
                    max_width = max(max_width, visual_len(cell.value))
            ws.column_dimensions[letter].width = max(min_w, min(max_width + pad, max_w))

def sanitize(name: str) -> str:
    return re.sub(r'[\\/*?:\[\]]', "_", str(name))[:31] or "BJ"

@st.cache_data(show_spinner=False, persist=False, ttl=0, max_entries=10)
def read_any_table(uploaded_file, sheet: str | int | None):
    name = (uploaded_file.name or "").lower()
    if name.endswith(".xlsx"):
        return pd.read_excel(uploaded_file, sheet_name=(sheet if str(sheet).strip() else 0))
    raw = uploaded_file.read(); uploaded_file.seek(0)
    for enc in ["utf-8", "utf-8-sig", "cp949", "euc-kr"]:
        try:
            text = raw.decode(enc)
            try:
                dialect = csv.Sniffer().sniff(text[:4000], delimiters=[",", "\t", ";", "|"])
                sep = dialect.delimiter
            except Exception:
                sep = ","
            return pd.read_csv(io.StringIO(text), sep=sep)
        except Exception:
            continue
    raise ValueError("CSV 인코딩/구분자 해석 실패")

# -------------------- core --------------------
def preprocess(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).strip() for c in df.columns]
    col_bj    = next((c for c in df.columns if c == "참여BJ"), None)
    col_heart = next((c for c in df.columns if c == "후원하트"), None)
    col_mix   = next((c for c in df.columns if c == "후원 아이디(닉네임)"), None)
    if not (col_bj and col_heart and col_mix):
        raise ValueError(f"필수 컬럼 누락: 참여BJ={col_bj}, 후원하트={col_heart}, 후원 아이디(닉네임)={col_mix}")

    df[col_bj] = df[col_bj].astype(str).str.strip()
    df[col_heart] = df[col_heart].astype(str).str.replace(",", "", regex=False)
    df[col_heart] = pd.to_numeric(df[col_heart], errors="coerce").fillna(0).astype(int)
    df[col_mix] = df[col_mix].astype(str).str.strip()

    sp = df[col_mix].str.extract(r'^\s*(?P<ID>[^()]+?)(?:\((?P<NICK>.*)\))?\s*$')
    df["ID"] = (sp["ID"].fillna("")
                .str.replace("\u200b","",regex=False)
                .str.replace("\ufeff","",regex=False)
                .str.replace("＠","@",regex=False)
                .str.strip())
    df["닉네임"] = sp["NICK"].fillna("").str.strip()

    base = (
        df.groupby([col_bj, "ID", "닉네임"], as_index=False)[col_heart]
          .sum()
          .rename(columns={col_bj:"참여BJ", col_heart:"후원하트"})
    )
    return base

def _xlsx_bytes_from_df(writer_fn) -> bytes:
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as w:
        writer_fn(w)
    bio.seek(0)
    wb = load_workbook(bio)
    autosize_columns(wb)  # 개선된 자동열폭 사용
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()

def make_bj_excel(bj_name: str, sub_df: pd.DataFrame, admin: bool) -> bytes:
    """admin=True: A~E열(구분/합계 포함) / admin=False: A~C열만"""
    sub = sub_df.copy()
    sub["is_aff"] = sub["ID"].str.contains("@")
    gen = sub[~sub["is_aff"]].sort_values("후원하트", ascending=False)[["ID","닉네임","후원하트"]].copy()
    aff = sub[ sub["is_aff"]].sort_values("후원하트", ascending=False)[["ID","닉네임","후원하트"]].copy()
    gsum = int(gen["후원하트"].sum()) if not gen.empty else 0
    asum = int(aff["후원하트"].sum()) if not aff.empty else 0
    total = gsum + asum
    sheet = sanitize(bj_name)

    def _write(w):
        # 1행: B=참여BJ, C=총합
        if admin:
            row1 = pd.DataFrame([[ "", bj_name, total, "", "" ]],
                                columns=["ID","닉네임","후원하트","구분","합계"])
        else:
            row1 = pd.DataFrame([[ "", bj_name, total ]],
                                columns=["ID","닉네임","후원하트"])
        row1.to_excel(w, sheet_name=sheet, index=False, header=False, startrow=0)

        # 2행: 헤더
        if admin:
            pd.DataFrame(columns=["ID","닉네임","후원하트","구분","합계"]).to_excel(
                w, sheet_name=sheet, index=False, startrow=1
            )
        else:
            pd.DataFrame(columns=["ID","닉네임","후원하트"]).to_excel(
                w, sheet_name=sheet, index=False, startrow=1
            )

        row = 2

        # 일반 블록
        if not gen.empty:
            blk = gen.copy()
            if admin:
                blk["구분"] = ""; blk["합계"] = ""
                blk.iloc[0, blk.columns.get_loc("구분")] = "일반하트"
                blk.iloc[0, blk.columns.get_loc("합계")] = gsum
            blk.to_excel(w, sheet_name=sheet, index=False, header=False, startrow=row)
            row += len(blk)

        # 제휴 블록
        if not aff.empty:
            blk = aff.copy()
            if admin:
                blk["구분"] = ""; blk["합계"] = ""
                blk.iloc[0, blk.columns.get_loc("구분")] = "제휴하트"
                blk.iloc[0, blk.columns.get_loc("합계")] = asum
            blk.to_excel(w, sheet_name=sheet, index=False, header=False, startrow=row)

    return _xlsx_bytes_from_df(_write)

def build_file_sets(base: pd.DataFrame):
    summary = base.groupby("참여BJ", as_index=False)["후원하트"].sum().sort_values("후원하트", ascending=False)

    def make_summary_bytes() -> bytes:
        return _xlsx_bytes_from_df(lambda w: summary.to_excel(w, sheet_name="요약", index=False))

    def pack_zip(files: dict[str, bytes]) -> bytes:
        zbio = io.BytesIO()
        with zipfile.ZipFile(zbio, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for fname, data in files.items():
                zf.writestr(fname, data)
        zbio.seek(0); return zbio.getvalue()

    # 관리자용
    admin_files: dict[str, bytes] = {"요약.xlsx": make_summary_bytes()}
    for bj in summary["참여BJ"]:
        sub = base[base["참여BJ"] == bj][["ID","닉네임","후원하트"]]
        admin_files[f"{sanitize(bj)}.xlsx"] = make_bj_excel(str(bj), sub, admin=True)
    admin_zip = pack_zip(admin_files)

    # BJ용
    bj_files: dict[str, bytes] = {"요약.xlsx": make_summary_bytes()}
    for bj in summary["참여BJ"]:
        sub = base[base["참여BJ"] == bj][["ID","닉네임","후원하트"]]
        bj_files[f"{sanitize(bj)}.xlsx"] = make_bj_excel(str(bj), sub, admin=False)
    bj_zip = pack_zip(bj_files)

    return (admin_files, admin_zip), (bj_files, bj_zip)

# -------------------- run --------------------
if uploaded:
    try:
        df_in = read_any_table(uploaded, sheet_name if uploaded.name.lower().endswith(".xlsx") else None)
        base = preprocess(df_in)
        (admin_files, admin_zip), (bj_files, bj_zip) = build_file_sets(base)

        left, right = st.columns(2, gap="large")

        with left:
            st.subheader("관리자용 (구분/합계 포함)")
            st.download_button("📦 관리자용 ZIP 다운로드", data=admin_zip,
                               file_name="BJ별_관리자용.zip", mime="application/zip",
                               use_container_width=True, key="zip-admin")
            st.divider(); st.markdown("**개별 파일 다운로드**")
            for i, (fname, data) in enumerate(admin_files.items()):
                st.download_button(f"⬇️ {fname}", data=data,
                                   file_name=fname,
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key=f"admin-{i}-{fname}")

        with right:
            st.subheader("BJ용 (D/E 없음)")
            st.download_button("📦 BJ용 ZIP 다운로드", data=bj_zip,
                               file_name="BJ별_BJ용.zip", mime="application/zip",
                               use_container_width=True, key="zip-bj")
            st.divider(); st.markdown("**개별 파일 다운로드**")
            for i, (fname, data) in enumerate(bj_files.items()):
                st.download_button(f"⬇️ {fname}", data=data,
                                   file_name=fname,
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key=f"bj-{i}-{fname}")

    except Exception as e:
        st.error(f"오류: {e}")
    finally:
        try:
            uploaded.close()
        except Exception:
            pass
        uploaded = None
else:
    st.info("CSV/XLSX 파일을 업로드하면 좌(관리자용), 우(BJ용)으로 각각 내려받을 수 있습니다. 모든 파일은 서버에 저장하지 않습니다.")
